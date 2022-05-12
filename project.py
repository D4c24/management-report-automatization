from socket import ALG_OP_VERIFY
from unicodedata import name
import requests, json, string, pandas as pd
from pprint import pprint
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import PieChart3D, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta

url_cloud = 'API url from the data source'
key_cloud = 'Key from the API'

##Turn string date object into deltatime
def to_deltatime(data):
  result = []
  for x in range(len(data)):
    if (none == data[x]):
      result.insert(x, 'N/A')
    else:
      date_select = datetime.strptime(data[x], '%Y-%m-%dT%H:%M:S.%f+0000')
      delta = timedelta(days=1)
      target_date = date_select + delta
      result.insert(x, target_date)
  return(result)

##Function to get the difference between two dates in a timedelta object
def diff_time(data1, data2):
  result = []
  for x in range(len(data2)):
    cond = isinstance(data2[x], str)
    if cond is True:
      result.insert(x, 'N/A')
    else:
      date = data2[x] - data1[x]
      result.insert(x, date)
  return(result)


##Function to get data from the Jira API
def Jira_api(url, token_key):
  header = {'Authorization': token_key, 'Content-Type': 'application/json'}
  jqltask = 'jql=project = test AND issuetyep = testIncident AND labels = sec AND created >= -30d'
  resp = (requests.get(url+jqltask, headers=header, verify=False)).text
  data = json.loads(resp)['issues']
  key = [i['key'] for i in data]
  summary, status, priority, reporter, created, updated, lastViewed, resolved, elapsedTime, elapsedTimeMillis, custom = [], [], [], [], [], [], [], [], [], [], []

  for i in range(len(data)):
    expand = data[i]
    fields = expand['fields']
    summary += [fields['summary']]
    dict_status = fields['status']
    status += [dict_status['name']]
    dict_priority = fields['priority']
    priority += [dict_priority['name']]
    dict_reporter = fields['reporter']
    reporter += [dict_reporter['name']]
    created += [fields['created']]
    updated += [fields['updated']]
    lastViewed =+ [fields['lastViewed']]
    resolved += [fields['resolutiondate']]
    custom = fields['customfield_10217']
    completedcycle = custom['completedCycles']

    if len(completedcycle) == 0:
      ongoingCycle = custom['ongoingCycle']
      elapsed = ongoingCycle['elapsedTime']
      elapsedTime.insert(i, elapsed['friendly'])
      elapsedTimeMillis.insert(i, int(elapsed['millis']))
    else:
      cycle = completedcycle[0]
      elapsed = cycle['elapsedTime']
      elapsedTime.insert(i, elapsed['friendly'])
      elapsedTimeMillis.insert(i, int(elapsed['millis']))

    created_data = to_deltatime(created)
    updated_data = to_deltatime(updated)
    lastV_data = to_deltatime(resolved)
    resolved_data = to_deltatime(resolved)
    time_to_resolve = diff_time(created_data, updated_data)

    columName = ['summary', 'key', 'status', 'priority', 'reporter', 'created', 'updated', 'lastViewed', 'resolved', 'elapsedTime', 'elapsedTimeMillis', 'to_resolve']
    list_of_Lists = [summary, key, status, priority, reporter, created, updated, lastViewed, resolved, elapsedTime, elapsedTimeMillis, time_to_resolve]

    df = pd.DataFrame()

    for x in range(len(list_of_Lists)):
      df.insert(x, columName[x], list_of_Lists[x])

    df['Today'] = pd.to_datetime('today')

    return(df)

def pie_char3D(ws, ref, lmin_col, min_row, dmin_col, tittle, dest_cel):
  count = len(ref.to_dict())
  max_row = min_row + (count - 1)
  pie = PieChart3D()
  labels = Reference(ws, min_col=lmin_col, min_row=min_row, max_row=max_row)
  data = Reference(ws, min_col=dmin_col, min_row=(min_row-1), max_row=max_row)
  pie.add_data(data, titles_from_data=True)
  pie.title = title
  ws.add_chart(pie, dest_cel)

def write_to_xls(data, ws, row):
  datadict = data.to_dict()
  keys = list(datadict.keys())
  values = list(datadict.values())
  cond = isinstance(values[0], dict)
  def write(listdata, rowVal, i, r):
    while (r + 1) <= len(listdata):
      ws.cell(column=i, row=(rowVal + r), value=listdata[r])
      r += 1
    if cond is True:
      for x in range(len(values)):
        ws.cell(column=(x + 1), row=(row - 1), value=keys[x])
        write(values[x], row, (x + 1), 0)
    else:
      write(keys, row, 1, 0)
      write(values, row, 2, 0)

def createTable(headings, data, ws, row, tName):
  data = data.values.tolist()
  alphString = string.ascii_uppercase
  alphList = list(alphString)
  reference = str(alphList[0] + str(row - 1) + ':' + alphList[len(headings) - 1] + str(row + len(data) - 1))
  def write(listdata, rowVal, i, r):
    while (r + 1) <= len(lsitdata):
      ws.cell(column=(i + r), row=rowVal, value=listdata[r])
      r += 1
  write(headings, (row - 1), 1, 0)
  for x in range(len(data)):
    write(data[x], (row + x), 1, 0)
  tab = Table(displayName = tName, ref = reference)
  style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=True, 
    showLastColumn=True, showRowStripes=True, showColumnStripes=True)
  tab.tableStyleInfo = style
  ws.add_table(tab)

def run():
  wb = Workbook()
  ws = wb.active
  ws1 = wb.create_sheet('charts', 0)
  ws2 = wb.create_sheet('B_cloud', 1)
  df = Jira_api(url_cloud, key_cloud)

  for r in dataframe_to_rows(df, index=True, header=True):
    ws2.append(r)

  wb.save('chart_data.xlsx')

  ##Pie chart tickets per status
  status = df.priority.value_counts()
  write_to_xls(status, ws1, 2)
  pie_char3D(ws1, priority, 1, 2, 2, 'Tickets per Priority', 'J2')

  ##Pie chart Tickets per priority
  priority = df.priority.value_counts()
  write_to_xls(priority, ws1, 18)
  pie_char3D(ws1, priority, 1, 18, 2, 'Ticket per Priority', 'J18')

  ##Top Critical Incidents
  namecolum = ['key', 'priority', 'status', 'created']
  top_5 = df[namecolum].sort_values(by='priority', ascending=True).head(5)
  createTable(namecolum, top_5, ws1, 35, 'Top_critical_incidents')

  ##Time to resolve - Average time VS count of tickets
  colname = ['states', 'priority', 'cxreated', 'resolved', 'to_resolve']
  priorList = ['Critical', 'High', 'Medium', 'Low']
  avrg_data, countlist = [], []
  for x in range(len(priorlist)):
    coldata = df[colname][(df.status=='Resolved') | (df.status=='Closed')]
    coldata = coldata[(coldata.priority==priorList[x])]
    if (coldata.priority.size) == 0:
      avrg_data.insert(x, 0)
      countlist.insert(x, 0)
    else:
      avrg_data.insert(x, (sum(coldata['to_resolve'], timedelta()) / len(coldata['to_resolve'])) / timedelta(days=1))
      countlist.insert(x, (coldata.priority.size))

  avrg_df = pd.DataFrame({'priority': priorList, 'Avrg_time(Days)': avrg_data, 'count(Tickets)': countlist})
  write_to_xls(avrg_df, ws1, 47)

  wb.save('chart_data.xlsx')
  
if __name__ == '__main__':
  run()
